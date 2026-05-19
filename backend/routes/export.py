"""
ComplAI — Export routes
Generate Busy Excel and TallyPrime XML from extracted invoices/bank statements.

INVOICE EXPORTS
  POST /api/export/busy         → .xlsx (3 sheets: Purchase Register, Busy Import v21, Busy Import v17-18)
  POST /api/export/tally        → .xml  (TallyPrime or ERP9 purchase vouchers)

BANK EXPORTS
  POST /api/export/bank         → .xlsx (4 sheets: Transactions, Summary, Busy Bank Import, Busy Bank Recon)
  POST /api/export/tally-bank   → .xml  (TallyPrime or ERP9 Receipt/Payment vouchers per bank transaction)

BANK SHEET GUIDE
  "Transactions"     — Full reference with match status (internal use)
  "Summary"          — Opening/closing balance, totals
  "Busy Bank Import" — Administration → Import Data → Vouchers (one voucher per row)
  "Busy Bank Recon"  — Administration → Import Data → Bank Statement (reconciliation, 5-column format)
"""

import io
import re
import xml.etree.ElementTree as ET
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from models.db import BankTransaction, Firm, Invoice, get_db
from routes.auth import get_current_firm

router = APIRouter()


# ── Request schemas ─────────────────────────────────────────

class ExportRequest(BaseModel):
    invoice_ids:   Optional[List[int]] = None
    client_id:     Optional[int]       = None
    month_year:    Optional[str]       = None   # "2025-10"
    tally_version: Optional[str]       = "prime"  # "prime" | "erp9"


class BankExportRequest(BaseModel):
    client_id:  int
    month_year: str


# ── Helpers ─────────────────────────────────────────────────

def _get_invoices(req: ExportRequest, firm: Firm, db: Session) -> List[Invoice]:
    """Resolve export target: either explicit IDs or client+month filter."""
    q = db.query(Invoice).join(Invoice.document).filter(
        Invoice.document.has(firm_id=firm.id)
    )
    if req.invoice_ids:
        q = q.filter(Invoice.id.in_(req.invoice_ids))
    elif req.client_id and req.month_year:
        q = q.filter(
            Invoice.client_id == req.client_id,
            Invoice.document.has(month_year=req.month_year),
        )
    else:
        raise HTTPException(status_code=400, detail="Provide invoice_ids OR client_id+month_year")
    return q.all()


def _fmt_date_ddmmyyyy(d: Optional[date]) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _fmt_date_tally(d: Optional[date]) -> str:
    """TallyPrime expects YYYYMMDD."""
    return d.strftime("%Y%m%d") if d else ""


# ── Shared styling helper ───────────────────────────────────

def _apply_header_style(ws, headers, header_fill, header_font, center=True):
    """Write headers with blue fill + white bold text. Returns nothing."""
    from openpyxl.styles import Alignment
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, headers):
    """Set column widths based on content."""
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# ── Busy Excel export ───────────────────────────────────────
#
# The workbook has THREE sheets:
#
#   Sheet 1 — "Purchase Register"  (CA reference / review)
#   Sheet 2 — "Busy Import v21"    (directly importable, Busy 19/20/21)
#   Sheet 3 — "Busy Import v17-18" (older Busy versions fallback)
#
# HOW TO IMPORT INTO BUSY (v19 / v20 / v21 — same steps):
#   1. Open Busy → Administration → Import Data → Vouchers
#   2. Select voucher type: Purchase
#   3. Browse to this Excel file, select sheet "Busy Import v21"
#   4. Busy shows a column-mapping screen — columns already match field names
#   5. Click Import
#   ⚠ Party names must exist in Busy Masters before import.
#      If a party doesn't exist, Busy either skips or auto-creates (depends on settings).
#
# HOW TO IMPORT INTO BUSY v17/v18:
#   Same steps but use sheet "Busy Import v17-18" — older format without GST %.

@router.post("/busy")
def export_busy(
    req:  ExportRequest,
    firm: Firm    = Depends(get_current_firm),
    db:   Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    invoices = _get_invoices(req, firm, db)
    if not invoices:
        raise HTTPException(status_code=404, detail={"error": "No invoices found", "error_code": "EXPORT_FAILED"})

    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="2E75B6")
    hdr_font  = Font(color="FFFFFF", bold=True)
    rev_fill  = PatternFill("solid", fgColor="FFF2CC")   # yellow = needs review
    grey_fill = PatternFill("solid", fgColor="E8F0FE")   # light blue = import sheet header

    # ═══════════════════════════════════════════════════════
    # SHEET 1 — Purchase Register (CA review / reference)
    # ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Purchase Register"

    reg_headers = [
        "Date", "Vendor Name", "Vendor GSTIN", "Invoice No", "Invoice Date",
        "Taxable Value", "CGST", "SGST", "IGST", "Total GST", "Invoice Total",
        "Transaction Type", "HSN/SAC", "Confidence %", "Needs Review", "Issues",
    ]
    _apply_header_style(ws1, reg_headers, hdr_fill, hdr_font)

    for inv in invoices:
        hsn_list = list({li.get("hsn_sac", "") for li in (inv.line_items or []) if li.get("hsn_sac")})
        row = [
            _fmt_date_ddmmyyyy(inv.invoice_date),
            inv.vendor_name or "",
            inv.vendor_gstin or "",
            inv.invoice_number or "",
            _fmt_date_ddmmyyyy(inv.invoice_date),
            inv.taxable_value or 0,
            inv.cgst or 0,
            inv.sgst or 0,
            inv.igst or 0,
            inv.total_gst or 0,
            inv.invoice_total or 0,
            (inv.transaction_type.value if inv.transaction_type else ""),
            ", ".join(filter(None, hsn_list)),
            f"{inv.combined_confidence * 100:.0f}%",
            "Yes" if inv.needs_review else "No",
            "; ".join(inv.issues or []),
        ]
        ws1.append(row)
        if inv.needs_review:
            for c in range(1, len(reg_headers) + 1):
                ws1.cell(row=ws1.max_row, column=c).fill = rev_fill

    _autowidth(ws1, reg_headers)

    # ═══════════════════════════════════════════════════════
    # SHEET 2 — Busy Import v19/20/21
    #
    # Column names match Busy's internal field labels exactly.
    # Busy's import wizard uses column-name matching — these
    # names auto-map without manual mapping on v19+.
    #
    # One row = one voucher (account-based, not item-based).
    # GST % columns are required by v19+ for GST ledger auto-fill.
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Busy Import v21")

    busy_v21_headers = [
        "Date",           # DD/MM/YYYY — voucher date
        "Vch No.",        # leave blank → Busy auto-assigns
        "Ref No.",        # supplier's invoice number
        "Ref Date",       # supplier's invoice date DD/MM/YYYY
        "Party Name",     # vendor/supplier name (must exist in Busy Masters)
        "Party GSTIN",    # vendor GSTIN
        "Place of Supply",# state name e.g. "Maharashtra"
        "Purchase A/c",   # purchase ledger name in Busy e.g. "Purchase"
        "Taxable Amount", # taxable value (number, no ₹)
        "CGST %",         # e.g. 9 (not 9%)
        "CGST Amount",    # e.g. 180.00
        "SGST %",         # e.g. 9
        "SGST Amount",    # e.g. 180.00
        "IGST %",         # e.g. 18 (blank for intrastate)
        "IGST Amount",    # e.g. 360.00 (blank for intrastate)
        "Cess Amount",    # 0 for most goods
        "Round Off",      # rounding difference if any
        "Net Amount",     # final invoice total (number, no ₹)
        "Narration",      # optional note
    ]
    _apply_header_style(ws2, busy_v21_headers, PatternFill("solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True))

    # Add an instruction row in row 2 (italic, grey)
    ws2.insert_rows(2)
    ws2.cell(row=2, column=1).value = (
        "⚠ Party Name must exist in Busy Masters. "
        "Import via: Administration → Import Data → Vouchers → Purchase → select this file → sheet 'Busy Import v21'"
    )
    ws2.cell(row=2, column=1).font = Font(italic=True, color="595959", size=9)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(busy_v21_headers))

    # State code → name map for Place of Supply
    STATE_NAMES = {
        "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
        "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
        "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
        "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
        "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
        "24":"Gujarat","25":"Daman & Diu","26":"Dadra & Nagar Haveli","27":"Maharashtra",
        "28":"Andhra Pradesh","29":"Karnataka","30":"Goa","31":"Lakshadweep",
        "32":"Kerala","33":"Tamil Nadu","34":"Puducherry","35":"Andaman & Nicobar Islands",
        "36":"Telangana","37":"Andhra Pradesh (New)","38":"Ladakh",
    }

    for inv in invoices:
        # Derive GST rates from amounts + taxable value
        taxable = inv.taxable_value or 0
        cgst_amt = inv.cgst or 0
        sgst_amt = inv.sgst or 0
        igst_amt = inv.igst or 0
        total    = inv.invoice_total or 0

        cgst_pct = round(cgst_amt / taxable * 100, 1) if taxable else 0
        sgst_pct = round(sgst_amt / taxable * 100, 1) if taxable else 0
        igst_pct = round(igst_amt / taxable * 100, 1) if taxable else 0

        # Place of supply from vendor GSTIN state code
        state_code = (inv.vendor_gstin or "")[:2]
        place_of_supply = STATE_NAMES.get(state_code, "")

        # Round-off: difference between parts and total
        parts_sum = round(taxable + cgst_amt + sgst_amt + igst_amt, 2)
        round_off = round(total - parts_sum, 2) if total else 0

        is_intrastate = (not inv.transaction_type) or inv.transaction_type.value == "intrastate"

        ws2.append([
            _fmt_date_ddmmyyyy(inv.invoice_date),  # Date
            "",                                      # Vch No. — let Busy assign
            inv.invoice_number or "",               # Ref No.
            _fmt_date_ddmmyyyy(inv.invoice_date),  # Ref Date
            inv.vendor_name or "",                  # Party Name
            inv.vendor_gstin or "",                 # Party GSTIN
            place_of_supply,                        # Place of Supply
            "Purchase",                             # Purchase A/c
            taxable,                                # Taxable Amount
            cgst_pct if is_intrastate else 0,       # CGST %
            cgst_amt if is_intrastate else 0,       # CGST Amount
            sgst_pct if is_intrastate else 0,       # SGST %
            sgst_amt if is_intrastate else 0,       # SGST Amount
            igst_pct if not is_intrastate else 0,   # IGST %
            igst_amt if not is_intrastate else 0,   # IGST Amount
            0,                                      # Cess Amount
            round_off,                              # Round Off
            total,                                  # Net Amount
            f"Invoice {inv.invoice_number or ''} from {inv.vendor_name or ''}",  # Narration
        ])
        # Yellow for needs-review rows
        if inv.needs_review:
            for c in range(1, len(busy_v21_headers) + 1):
                ws2.cell(row=ws2.max_row, column=c).fill = rev_fill

    _autowidth(ws2, busy_v21_headers)

    # ═══════════════════════════════════════════════════════
    # SHEET 3 — Busy Import v17/v18  (older Busy versions)
    #
    # Pre-GST era Busy (v17/18) doesn't have GST % columns.
    # Uses simpler voucher format.
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Busy Import v17-18")

    busy_old_headers = [
        "Date",           # DD/MM/YYYY
        "Vch No.",        # blank = auto
        "Ref No.",        # invoice number
        "Ref Date",       # invoice date
        "Party Name",     # vendor name
        "Gross Amount",   # invoice total
        "Discount",       # 0 for most
        "Net Amount",     # same as gross usually
        "Narration",      # optional
    ]
    _apply_header_style(ws3, busy_old_headers, PatternFill("solid", fgColor="833C00"), Font(color="FFFFFF", bold=True))

    ws3.insert_rows(2)
    ws3.cell(row=2, column=1).value = (
        "⚠ For Busy v17/v18 only. GST columns not supported in these versions — enter GST manually after import."
    )
    ws3.cell(row=2, column=1).font = Font(italic=True, color="595959", size=9)
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(busy_old_headers))

    for inv in invoices:
        total = inv.invoice_total or 0
        ws3.append([
            _fmt_date_ddmmyyyy(inv.invoice_date),
            "",
            inv.invoice_number or "",
            _fmt_date_ddmmyyyy(inv.invoice_date),
            inv.vendor_name or "",
            total,
            0,
            total,
            f"Invoice {inv.invoice_number or ''} from {inv.vendor_name or ''}",
        ])

    _autowidth(ws3, busy_old_headers)

    # ── Stream the file ────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"busy_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Tally XML export ────────────────────────────────────────

@router.post("/tally")
def export_tally(
    req:  ExportRequest,
    firm: Firm    = Depends(get_current_firm),
    db:   Session = Depends(get_db),
):
    """
    Generate Tally-compatible purchase voucher XML.

    tally_version:
      "prime"  → TallyPrime 2.x / 3.x / 4.x  (ALLLEDGERENTRIES.LIST, full GST tags)
      "erp9"   → Tally ERP 9                   (LEDGERENTRIES.LIST, simpler schema)

    Import via:
      TallyPrime : Gateway of Tally → Import → Data → select XML file
      Tally ERP9 : Gateway of Tally → Import Data → Vouchers → select XML file
    """
    invoices = _get_invoices(req, firm, db)
    if not invoices:
        raise HTTPException(status_code=404, detail={"error": "No invoices found", "error_code": "EXPORT_FAILED"})

    version = (req.tally_version or "prime").lower()
    is_erp9 = version == "erp9"
    # ERP9 uses LEDGERENTRIES.LIST; TallyPrime uses ALLLEDGERENTRIES.LIST
    ENTRY_TAG = "LEDGERENTRIES.LIST" if is_erp9 else "ALLLEDGERENTRIES.LIST"

    # Build XML tree
    envelope = ET.Element("ENVELOPE")

    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")

    req_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(req_desc, "REPORTNAME").text = "Vouchers"

    req_data = ET.SubElement(import_data, "REQUESTDATA")

    skipped = []
    for inv in invoices:
        # ── Resolve amounts, deriving missing values ───────────────────────────
        total   = float(inv.grand_total or inv.invoice_total or 0)
        cgst    = float(inv.cgst  or 0)
        sgst    = float(inv.sgst  or 0)
        igst    = float(inv.igst  or 0)
        gst_sum = round(cgst + sgst + igst, 2)

        # If taxable_value missing, back-calculate from total - gst
        taxable = float(inv.taxable_value or 0)
        if taxable == 0 and total > 0:
            taxable = round(total - gst_sum, 2)

        # If total missing, build from parts
        if total == 0 and taxable > 0:
            total = round(taxable + gst_sum, 2)

        # Skip entirely if we have nothing at all
        if total == 0 and taxable == 0:
            skipped.append(inv.invoice_number or str(inv.id))
            continue

        # Check balance: credit (vendor) must equal sum of debits
        balance_check = round(total - taxable - gst_sum, 2)
        # If still off by small rounding, absorb into purchase amount
        adjusted_taxable = round(taxable + balance_check, 2)

        vendor_name = inv.vendor_name or "Unknown Vendor"
        date_str    = _fmt_date_tally(inv.invoice_date)
        inv_number  = inv.invoice_number or str(inv.id)

        is_intrastate = (
            inv.transaction_type is None  # default to intrastate when unknown
            or inv.transaction_type.value == "intrastate"
        )

        # ── Build TALLYMESSAGE ─────────────────────────────────────────────────
        msg = ET.SubElement(req_data, "TALLYMESSAGE")
        msg.set("xmlns:UDF", "TallyUDF")

        voucher = ET.SubElement(msg, "VOUCHER")
        voucher.set("REMOTEID", f"{inv.id}")   # DB id is always unique
        voucher.set("VCHTYPE", "Purchase")
        voucher.set("ACTION", "Create")

        ET.SubElement(voucher, "DATE").text            = date_str
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Purchase"
        ET.SubElement(voucher, "VOUCHERNUMBER").text   = inv_number
        ET.SubElement(voucher, "NARRATION").text = (
            f"Invoice {inv_number} from {vendor_name}"
            + (f" | GSTIN: {inv.vendor_gstin}" if inv.vendor_gstin else "")
        )

        # TallyPrime-only tags (ERP9 doesn't support these)
        if not is_erp9:
            ET.SubElement(voucher, "EFFECTIVEDATE").text   = date_str
            ET.SubElement(voucher, "PARTYLEDGERNAME").text = vendor_name
            ET.SubElement(voucher, "ISINVOICE").text       = "Yes"

        def _ledger_entry(parent, ledger, is_debit: bool, amount: float):
            """Add a ledger entry. Tally convention: credit=positive, debit=negative."""
            e = ET.SubElement(parent, ENTRY_TAG)
            ET.SubElement(e, "LEDGERNAME").text       = ledger
            ET.SubElement(e, "ISDEEMEDPOSITIVE").text = "Yes" if is_debit else "No"
            ET.SubElement(e, "AMOUNT").text           = f"-{amount:.2f}" if is_debit else f"{amount:.2f}"

        # ── Credit: Vendor / Sundry Creditor ──────────────────────────────────
        _ledger_entry(voucher, vendor_name, is_debit=False, amount=total)

        # ── Debit: Purchase Account ────────────────────────────────────────────
        _ledger_entry(voucher, "Purchase Account", is_debit=True, amount=adjusted_taxable)

        # ── Debit: GST Input ledgers ───────────────────────────────────────────
        if is_intrastate:
            if cgst > 0:
                _ledger_entry(voucher, "CGST Input", is_debit=True, amount=cgst)
            if sgst > 0:
                _ledger_entry(voucher, "SGST Input", is_debit=True, amount=sgst)
        else:
            if igst > 0:
                _ledger_entry(voucher, "IGST Input", is_debit=True, amount=igst)

    # Add a comment listing any skipped invoices so the CA knows
    if skipped:
        comment = ET.Comment(f" Skipped (zero amounts, unfixable): {', '.join(skipped)} ")
        req_data.append(comment)

    xml_str = ET.tostring(envelope, encoding="unicode", xml_declaration=False)
    xml_bytes = ('<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str).encode("utf-8")

    ver_label = "erp9" if is_erp9 else "prime"
    filename = f"tally_{ver_label}_vouchers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Bank Excel export ───────────────────────────────────────

@router.post("/bank")
def export_bank(
    req:  BankExportRequest,
    firm: Firm    = Depends(get_current_firm),
    db:   Session = Depends(get_db),
):
    """
    Export bank transactions as Excel with two sheets:
      Sheet 1 "Transactions" — all rows with voucher type
      Sheet 2 "Summary"      — opening/closing balance, totals
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    txs = (
        db.query(BankTransaction)
        .filter(
            BankTransaction.firm_id == firm.id,
            BankTransaction.client_id == req.client_id,
            BankTransaction.month_year == req.month_year,
        )
        .order_by(BankTransaction.transaction_date)
        .all()
    )
    if not txs:
        raise HTTPException(status_code=404, detail={"error": "No transactions found", "error_code": "EXPORT_FAILED"})

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transactions ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Transactions"

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(color="FFFFFF", bold=True)
    unmatched_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = ["Date", "Narration", "Debit", "Credit", "Balance",
               "Voucher Type", "Ledger Match", "Confirmed"]
    ws1.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for tx in txs:
        row = [
            _fmt_date_ddmmyyyy(tx.transaction_date),
            tx.narration or "",
            tx.debit or "",
            tx.credit or "",
            tx.balance or "",
            tx.voucher_type.value if tx.voucher_type else "",
            tx.matched_ledger or "",
            "Yes" if tx.match_confirmed else "No",
        ]
        ws1.append(row)
        if not tx.matched_ledger:
            row_idx = ws1.max_row
            for col_idx in range(1, len(headers) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = unmatched_fill

    for col_idx, hdr in enumerate(headers, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(hdr) + 4, 15)

    # ── Sheet 2: Summary ───────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total_credits = sum(t.credit or 0 for t in txs)
    total_debits  = sum(t.debit  or 0 for t in txs)
    opening_bal   = txs[0].balance - (txs[0].credit or 0) + (txs[0].debit or 0) if txs else 0
    closing_bal   = txs[-1].balance if txs else 0
    unmatched_cnt = sum(1 for t in txs if not t.matched_ledger)

    summary_rows = [
        ("Period",             req.month_year),
        ("Opening Balance",    f"₹{opening_bal:,.2f}"),
        ("Total Credits",      f"₹{total_credits:,.2f}"),
        ("Total Debits",       f"₹{total_debits:,.2f}"),
        ("Closing Balance",    f"₹{closing_bal:,.2f}"),
        ("Total Transactions", len(txs)),
        ("Unmatched",          unmatched_cnt),
    ]
    for label, value in summary_rows:
        ws2.append([label, value])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    # ═══════════════════════════════════════════════════════
    # SHEET 3 — Busy Bank Voucher Import (v19/20/21)
    #
    # Busy imports bank entries as Receipt / Payment / Contra vouchers.
    # One row per transaction. Each row maps to one voucher in Busy.
    #
    # HOW TO IMPORT:
    #   Administration → Import Data → Vouchers → select Vch Type per batch
    #   (import Receipts first, then Payments, then Contra)
    #   OR: import all rows together — Busy reads "Vch Type" column to split.
    #
    # LEDGER NAMES:
    #   "Bank A/c" column → your bank account name in Busy (e.g. "SBI Current A/c")
    #   "Party/Ledger A/c" → the matched ledger / party name in Busy
    #   Unmatched rows have "*** UNMATCHED ***" — CA must fill before importing.
    # ═══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Busy Bank Import")

    bank_import_headers = [
        "Date",            # DD/MM/YYYY
        "Vch Type",        # Receipt / Payment / Contra / Journal
        "Vch No.",         # blank = Busy auto-assigns
        "Bank A/c",        # your bank ledger name in Busy
        "Party/Ledger A/c",# matched party or ledger name
        "Amount",          # always positive number
        "Narration",       # bank narration
        "Ref No.",         # reference / cheque number (from narration if found)
        "Import Status",   # READY / UNMATCHED (CA fills blanks before importing)
    ]
    _apply_header_style(ws3, bank_import_headers,
                        PatternFill("solid", fgColor="1F4E79"),
                        Font(color="FFFFFF", bold=True))

    ws3.insert_rows(2)
    ws3.cell(row=2, column=1).value = (
        "⚠ Fill 'Bank A/c' with your exact bank ledger name in Busy. "
        "Fix all 'UNMATCHED' rows in 'Party/Ledger A/c' before importing. "
        "Import: Administration → Import Data → Vouchers → select this file → sheet 'Busy Bank Import'"
    )
    ws3.cell(row=2, column=1).font = Font(italic=True, color="595959", size=9)
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(bank_import_headers))

    import re as _re
    unmatched_fill2 = PatternFill("solid", fgColor="FFE0E0")  # red tint = must fix
    ready_fill      = PatternFill("solid", fgColor="E2EFDA")  # green tint = ready

    for tx in txs:
        vch_type_map = {
            "receipt": "Receipt",
            "payment": "Payment",
            "contra":  "Contra",
            "journal": "Journal",
        }
        vch_type = vch_type_map.get(tx.voucher_type.value if tx.voucher_type else "", "Journal")

        # Amount is always positive in Busy — direction is determined by Vch Type
        amount = tx.credit if tx.credit else (tx.debit or 0)

        # Try to extract cheque/ref number from narration
        ref_match = _re.search(r'(?:chq|cheque|ref|neft|rtgs|utr)[:\s#]*([A-Z0-9]+)',
                               (tx.narration or ""), _re.IGNORECASE)
        ref_no = ref_match.group(1) if ref_match else ""

        is_matched = bool(tx.matched_ledger)
        party_ledger = tx.matched_ledger if is_matched else "*** UNMATCHED — fill before import ***"
        import_status = "READY" if is_matched else "UNMATCHED"

        ws3.append([
            _fmt_date_ddmmyyyy(tx.transaction_date),
            vch_type,
            "",              # Vch No — let Busy assign
            "Bank A/c",      # CA replaces with their exact bank ledger name
            party_ledger,
            amount,
            tx.narration or "",
            ref_no,
            import_status,
        ])

        row_idx = ws3.max_row
        fill = ready_fill if is_matched else unmatched_fill2
        for c in range(1, len(bank_import_headers) + 1):
            ws3.cell(row=row_idx, column=c).fill = fill

    _autowidth(ws3, bank_import_headers)

    # ═══════════════════════════════════════════════════════
    # SHEET 4 — Busy Bank Recon (Bank Statement Reconciliation)
    #
    # This is the RECONCILIATION import — not voucher import.
    # Busy uses this to reconcile your already-entered vouchers
    # against the actual bank statement.
    #
    # HOW TO IMPORT:
    #   Administration → Import Data → Bank Statement
    #   Select this file → sheet "Busy Bank Recon"
    #   Map columns: Date, Description, Withdrawal, Deposit, Balance
    #
    # FORMAT RULES:
    #   • Exactly 5 columns — no extra columns allowed
    #   • Date: DD/MM/YYYY (Busy also accepts YYYY-MM-DD)
    #   • Withdrawal: amount if money went OUT of bank (debit), else blank
    #   • Deposit:    amount if money came IN to bank (credit), else blank
    #   • Balance:    running balance after this transaction (closing balance)
    #   • All amounts are positive numbers (direction shown by column)
    # ═══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Busy Bank Recon")

    recon_headers = ["Date", "Description", "Withdrawal", "Deposit", "Balance"]
    _apply_header_style(ws4, recon_headers,
                        PatternFill("solid", fgColor="375623"),   # dark green header
                        Font(color="FFFFFF", bold=True))

    ws4.insert_rows(2)
    ws4.cell(row=2, column=1).value = (
        "⚠ Bank Statement Reconciliation import. "
        "Administration → Import Data → Bank Statement → select this file → sheet 'Busy Bank Recon'. "
        "Withdrawal = money out (debit); Deposit = money in (credit). Balance = running balance. "
        "Blank cells in Withdrawal/Deposit are intentional — do not fill 0."
    )
    ws4.cell(row=2, column=1).font = Font(italic=True, color="595959", size=9)
    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(recon_headers))

    credit_fill_recon = PatternFill("solid", fgColor="E2EFDA")   # green = deposit
    debit_fill_recon  = PatternFill("solid", fgColor="FCE4D6")   # orange = withdrawal

    for tx in txs:
        withdrawal = tx.debit  if tx.debit  else None   # money out — positive number
        deposit    = tx.credit if tx.credit else None   # money in  — positive number
        balance    = tx.balance if tx.balance is not None else ""

        ws4.append([
            _fmt_date_ddmmyyyy(tx.transaction_date),
            tx.narration or "",
            withdrawal,    # blank cell if None — Busy needs blank, not 0
            deposit,       # blank cell if None
            balance,
        ])

        row_idx = ws4.max_row
        fill = credit_fill_recon if deposit else debit_fill_recon
        for c in range(1, len(recon_headers) + 1):
            ws4.cell(row=row_idx, column=c).fill = fill

    _autowidth(ws4, recon_headers)

    # ── Stream ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bank_export_{req.client_id}_{req.month_year}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Tally Bank XML export ───────────────────────────────────

class TallyBankExportRequest(BaseModel):
    client_id:     int
    month_year:    str
    tally_version: Optional[str] = "prime"   # "prime" | "erp9"
    bank_ledger:   Optional[str] = "Bank Account"  # Name of your bank ledger in Tally


@router.post("/tally-bank")
def export_tally_bank(
    req:  TallyBankExportRequest,
    firm: Firm    = Depends(get_current_firm),
    db:   Session = Depends(get_db),
):
    """
    Generate Tally-compatible XML for bank transactions.

    Each transaction becomes a voucher:
      Credit (money IN)  → Receipt voucher
        Debit  Bank Account  (asset increases)
        Credit Party Ledger  (income / liability decreases)

      Debit  (money OUT) → Payment voucher
        Credit Bank Account  (asset decreases)
        Debit  Party Ledger  (expense / asset increases)

    tally_version:
      "prime" → TallyPrime 2.x / 3.x / 4.x  (ALLLEDGERENTRIES.LIST)
      "erp9"  → Tally ERP 9                   (LEDGERENTRIES.LIST)

    bank_ledger:
      The exact name of your bank account ledger in Tally
      e.g. "SBI Current A/c", "HDFC Bank", "State Bank of India"
      Defaults to "Bank Account" — CA must rename to match their Tally books.

    Import via:
      TallyPrime : Gateway of Tally → Import → Data → select XML file
      Tally ERP9 : Gateway of Tally → Import Data → Vouchers → select XML file

    Unmatched transactions use "Suspense Account" as counter-ledger.
    CA should create a "Suspense Account" ledger in Tally if it doesn't exist,
    then reclassify those vouchers manually after import.
    """
    txs = (
        db.query(BankTransaction)
        .filter(
            BankTransaction.firm_id    == firm.id,
            BankTransaction.client_id  == req.client_id,
            BankTransaction.month_year == req.month_year,
        )
        .order_by(BankTransaction.transaction_date)
        .all()
    )
    if not txs:
        raise HTTPException(
            status_code=404,
            detail={"error": "No transactions found", "error_code": "EXPORT_FAILED"}
        )

    version   = (req.tally_version or "prime").lower()
    is_erp9   = (version == "erp9")
    ENTRY_TAG = "LEDGERENTRIES.LIST" if is_erp9 else "ALLLEDGERENTRIES.LIST"

    bank_ledger = req.bank_ledger or "Bank Account"

    # ── Build XML tree ─────────────────────────────────────
    envelope = ET.Element("ENVELOPE")

    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

    body        = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")

    req_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(req_desc, "REPORTNAME").text = "Vouchers"

    req_data = ET.SubElement(import_data, "REQUESTDATA")

    def _ledger_entry(parent, ledger: str, is_debit: bool, amount: float):
        """
        Add a ledger entry to a voucher.
        Tally convention: credit = positive AMOUNT, debit = negative AMOUNT.
        ISDEEMEDPOSITIVE = Yes means the entry is a debit (increases the account).
        """
        e = ET.SubElement(parent, ENTRY_TAG)
        ET.SubElement(e, "LEDGERNAME").text       = ledger
        ET.SubElement(e, "ISDEEMEDPOSITIVE").text = "Yes" if is_debit else "No"
        ET.SubElement(e, "AMOUNT").text           = f"-{amount:.2f}" if is_debit else f"{amount:.2f}"

    skipped = []

    for tx in txs:
        is_credit = bool(tx.credit) and (tx.credit or 0) > 0
        is_debit  = bool(tx.debit)  and (tx.debit  or 0) > 0

        amount = float(tx.credit or tx.debit or 0)
        if amount <= 0:
            skipped.append(f"tx#{tx.id} (zero amount)")
            continue

        # Voucher type: Receipt = money in, Payment = money out
        vch_type = "Receipt" if is_credit else "Payment"

        # Counter-ledger: matched party, or Suspense for unmatched
        counter_ledger = tx.matched_ledger if tx.matched_ledger else "Suspense Account"

        date_str = _fmt_date_tally(tx.transaction_date)
        narration = (
            (tx.narration or "Bank transaction")
            + (f" | Ref: {tx.matched_ledger}" if tx.matched_ledger else " | ⚠ Unmatched — reclassify in Tally")
        )

        # ── Build TALLYMESSAGE ────────────────────────────
        msg = ET.SubElement(req_data, "TALLYMESSAGE")
        msg.set("xmlns:UDF", "TallyUDF")

        voucher = ET.SubElement(msg, "VOUCHER")
        voucher.set("REMOTEID", f"bank_{tx.id}")
        voucher.set("VCHTYPE", vch_type)
        voucher.set("ACTION", "Create")

        ET.SubElement(voucher, "DATE").text            = date_str
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = vch_type
        ET.SubElement(voucher, "NARRATION").text       = narration

        if not is_erp9:
            ET.SubElement(voucher, "EFFECTIVEDATE").text = date_str

        if is_credit:
            # Receipt: Bank A/c debited (asset ↑), Party/Income ledger credited
            _ledger_entry(voucher, bank_ledger,    is_debit=True,  amount=amount)
            _ledger_entry(voucher, counter_ledger, is_debit=False, amount=amount)
        else:
            # Payment: Bank A/c credited (asset ↓), Party/Expense ledger debited
            _ledger_entry(voucher, counter_ledger, is_debit=True,  amount=amount)
            _ledger_entry(voucher, bank_ledger,    is_debit=False, amount=amount)

    if skipped:
        req_data.append(ET.Comment(f" Skipped (zero/missing amount): {', '.join(skipped)} "))

    xml_str   = ET.tostring(envelope, encoding="unicode", xml_declaration=False)
    xml_bytes = ('<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str).encode("utf-8")

    ver_label = "erp9" if is_erp9 else "prime"
    filename  = f"tally_{ver_label}_bank_{req.client_id}_{req.month_year}.xml"
    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
